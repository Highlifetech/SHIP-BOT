"""
Packing lists in Off Menu's own template, built from the rows the team
already fills in.

A shipment in the inbound sheets isn't one row -- it's every row sharing a
tracking number. Tracking #, carrier and box count are typed once on the first
row and carried down; each row under it is a line: a product photo, a
description, a quantity shipped and a quantity ordered. That is a packing list
already. Nobody has to enter anything twice.

    1ZC228W90429665565  7Brew Coffee   Brewista Crew - Black   300 / 300
                                       Brewista Crew - Nat.    225 / 300   75 short
                                       Trucker Cap - Red       200 / 200

What comes out the other end is Off Menu's Inside_Box_Packing_List layout:
letterhead and logo, ship-to block, Date / Invoice # / PO Reference,
Carrier / DDP / Date Shipped, then # of Boxes . Description . Tracking . QTY
over Total Boxes and Total Quantity, and a Special Notes / Photos strip.

Three things this module knows how to do:

    gather(results, tracking)   the rows for one shipment, split into boxes
    flags(doc)                  every template field the sheet can't fill
    render(doc)                 the printable set: a label and a list per box
    workbook(doc)               the same thing as .xlsx, one sheet per box

The flags matter more than they look. The template has fields the inbound
sheets have no column for -- ship-to address, invoice number, date shipped --
so rather than printing a form with silent holes in it, every gap is named on
screen before anyone prints. Short quantities are flagged the same way.
"""

import base64
import html
import io
import logging
import os
from datetime import datetime, timedelta, timezone

logger = logging.getLogger(__name__)

COMPANY = "OFF MENU"
COMPANY_LINES = ["115 Monterey Bay Drive, Boynton Beach, FL 33426",
                 "www.byoffmenu.com"]

LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                         "assets", "off-menu-logo.png")
_logo_cache = {}


def logo_bytes():
    """The Off Menu mark.

    Prefers assets/off-menu-logo.png so the logo can be swapped by dropping
    in a new file, and falls back to the copy inlined in logo_data -- which
    is what actually ships, since a flat list of .py files can't land in the
    wrong folder the way an assets/ directory can.
    """
    if "raw" not in _logo_cache:
        raw = b""
        try:
            with open(LOGO_PATH, "rb") as f:
                raw = f.read()
        except OSError:
            try:
                import logo_data
                raw = base64.b64decode(logo_data.PNG_B64)
            except Exception as e:
                logger.warning("No logo available (%s)", e)
        _logo_cache["raw"] = raw
    return _logo_cache["raw"]


def logo_src():
    """A data: URI, so a printed page carries the logo with it."""
    if "src" not in _logo_cache:
        raw = logo_bytes()
        _logo_cache["src"] = ("data:image/png;base64," +
                              base64.b64encode(raw).decode()) if raw else ""
    return _logo_cache["src"]


def _mark(cls=""):
    src = logo_src()
    if src:
        return '<img class="mark %s" src="%s" alt="Off Menu">' % (cls, src)
    return '<div class="mark %s">OFF<br>MENU</div>' % cls

# The template's own colours, sampled from the sheet.
INK = "#1c1c1c"
BAND = "#7f8c7f"        # the grey-green header bands
BAND_XL = "FF7F8C7F"
STRIPE = "#e8e8e8"      # the alternating row fill
STRIPE_XL = "FFE8E8E8"
TOTALS = "#fbe4d0"      # the peach totals row
TOTALS_XL = "FFFBE4D0"
NOTES = "#1c1c1c"       # the black Special Notes strip


def _int(v):
    v = str(v or "").strip().replace(",", "")
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def _s(v):
    return str(v or "").strip()


def _now_et():
    return datetime.now(timezone(timedelta(hours=-4)))


# ---------------------------------------------------------------- gathering

def gather(results, tracking):
    """Every row that belongs to one tracking number, split into boxes.

    Box allocation is per-line when the sheet carries it and everything lands
    in box 1 otherwise -- flagged, not guessed silently.
    """
    want = _s(tracking).upper()
    rows = [r for r in (results or [])
            if _s(r.get("tracking_num")).upper() == want]
    if not rows:
        return None

    head = rows[0]
    items = []
    for r in rows:
        name = _s(r.get("product_name"))
        shipped = _int(r.get("qty_shipped"))
        ordered = _int(r.get("qty_expected"))
        # A row with neither a description nor a quantity is a carry-forward
        # continuation, not a line anyone packed.
        if not name and not shipped and not ordered:
            continue
        items.append({
            "description": name,
            "photo": _s(r.get("product_photo")),
            "order_num": _s(r.get("order_num")),
            "shipped": shipped,
            "ordered": ordered,
            "short": max(0, ordered - shipped),
            "box": _int(r.get("box_num")) or 0,
            "notes": _s(r.get("notes")),
        })

    box_count = _int(head.get("num_boxes"))
    allocated = any(i["box"] for i in items)
    if not allocated:
        for i in items:
            i["box"] = 1
    if not box_count:
        box_count = max([i["box"] for i in items] or [1])

    boxes = []
    for n in range(1, box_count + 1):
        contents = [i for i in items if i["box"] == n]
        boxes.append({
            "num": n,
            "of": box_count,
            "items": contents,
            "qty": sum(i["shipped"] for i in contents),
        })

    doc = {
        "tracking": _s(head.get("tracking_num")),
        "carrier": _s(head.get("carrier")).upper(),
        "client": _s(head.get("customer")) or _s(head.get("recipient")),
        "ship_to": _s(head.get("ship_to")) or _s(head.get("address")),
        "shipment_id": _s(head.get("shipment_id")),
        "invoice_num": _s(head.get("invoice_num")),
        "po_ref": _s(head.get("order_num")),
        "ddp": _s(head.get("ddp")) or _s(head.get("tariff_charge")),
        "date_shipped": _s(head.get("ship_date")),
        "tab": _s(head.get("tab")),
        "notes": _s(head.get("notes")),
        "items": items,
        "boxes": boxes,
        "box_count": box_count,
        "allocated": allocated,
        "shipped_total": sum(i["shipped"] for i in items),
        "ordered_total": sum(i["ordered"] for i in items),
        "generated": _now_et().strftime("%B %-d, %Y"),
    }
    doc["short_total"] = max(0, doc["ordered_total"] - doc["shipped_total"])
    return doc


# ----------------------------------------------------------------- flagging

def flags(doc):
    """Everything the template asks for that this shipment can't answer.

    Ordered worst-first: shortages, then blank header fields, then lines with
    holes in them. Each is {level, text}; 'stop' is a shipment nobody should
    print yet, 'warn' is a hole someone has to fill by hand.
    """
    out = []

    def add(level, text):
        out.append({"level": level, "text": text})

    if doc["short_total"]:
        add("stop", "%d units short of what was ordered (%d shipped of %d)"
            % (doc["short_total"], doc["shipped_total"], doc["ordered_total"]))
    if not doc["items"]:
        add("stop", "No line items on this shipment yet")
    if not doc["allocated"] and doc["box_count"] > 1:
        add("stop", "No box allocation - all %d lines are on Box 1 of %d"
            % (len(doc["items"]), doc["box_count"]))

    for label, value in (("Ship-to address", doc["ship_to"]),
                         ("Invoice #", doc["invoice_num"]),
                         ("PO Reference", doc["po_ref"]),
                         ("Carrier", doc["carrier"]),
                         ("DDP", doc["ddp"]),
                         ("Date Shipped", doc["date_shipped"]),
                         ("Tracking", doc["tracking"]),
                         ("Client", doc["client"])):
        if not value:
            add("warn", "%s is blank" % label)
    if not doc["box_count"]:
        add("warn", "# of Boxes is blank")

    missing_desc = sum(1 for i in doc["items"] if not i["description"])
    if missing_desc:
        add("warn", "%d line%s with no description" %
            (missing_desc, "" if missing_desc == 1 else "s"))
    no_qty = sum(1 for i in doc["items"] if not i["shipped"])
    if no_qty:
        add("warn", "%d line%s with no quantity shipped" %
            (no_qty, "" if no_qty == 1 else "s"))
    no_photo = sum(1 for i in doc["items"] if not i["photo"].startswith("http"))
    if no_photo:
        add("info", "%d line%s with no product photo" %
            (no_photo, "" if no_photo == 1 else "s"))

    empty = [b["num"] for b in doc["boxes"] if not b["items"]]
    if empty:
        add("warn", "Box%s %s empty" %
            ("es" if len(empty) > 1 else "",
             ", ".join(str(n) for n in empty) +
             (" are" if len(empty) > 1 else " is")))
    return out


# ---------------------------------------------------------------- rendering

def _head_block(doc, e):
    """Letterhead, ship-to and the Date / Invoice / PO panel."""
    def field(label, value):
        cls = "" if value else ' class="hole"'
        return ('<tr><th>%s</th><td%s>%s</td></tr>'
                % (e(label), cls, e(value) if value else "&mdash;"))
    ship = e(doc["ship_to"]).replace("\n", "<br>") if doc["ship_to"] else \
        '<span class="hole">&mdash;</span>'
    return """
  <div class="top">
    <div class="from">
      <b>%(company)s</b>%(colines)s
    </div>
    <div class="logo">%(logo)s</div>
  </div>
  <div class="mid">
    <div class="shipto">
      <div class="cap">SHIPPING ADDRESS</div>
      <div class="addr">%(who)s%(ship)s</div>
    </div>
    <table class="panel">%(fields)s</table>
  </div>""" % {
        "company": e(COMPANY),
        "colines": "".join("<div>%s</div>" % e(l) for l in COMPANY_LINES),
        "logo": _mark(),
        "who": ('<b>%s</b><br>' % e(doc["client"])) if doc["client"] else "",
        "ship": ship,
        "fields": (field("Date:", doc["date_shipped"] or doc["generated"])
                   + field("Invoice #:", doc["invoice_num"])
                   + field("PO Reference:", doc["po_ref"])),
    }


def _carrier_block(doc, e):
    def cell(v):
        return ('<td>%s</td>' % e(v)) if v else '<td class="hole">&mdash;</td>'
    return """
  <table class="carrier">
    <thead><tr><th>Carrier</th><th>DDP</th><th>Date Shipped</th></tr></thead>
    <tbody><tr>%s%s%s</tr></tbody>
  </table>""" % (cell(doc["carrier"]), cell(doc["ddp"]),
                 cell(doc["date_shipped"]))


def _sheet(doc, box, e):
    """One Inside_Box_Packing_List page."""
    rows = []
    for i, it in enumerate(box["items"]):
        desc = e(it["description"]) if it["description"] else \
            '<span class="hole">no description</span>'
        if it["order_num"]:
            desc += '<span class="ord">PO %s</span>' % e(it["order_num"])
        qty = ('<span class="short">%d of %d</span>' % (it["shipped"], it["ordered"])
               if it["short"] else str(it["shipped"]))
        rows.append('<tr><td class="c">%s</td><td>%s</td><td class="t">%s</td>'
                    '<td class="q">%s</td></tr>'
                    % ("1" if i == 0 else "", desc, e(doc["tracking"]) if i == 0 else "",
                       qty))
    while len(rows) < 8:      # keep the ruled look of the blank template
        rows.append('<tr class="blank"><td></td><td></td><td></td><td></td></tr>')

    photos = "".join(
        '<img src="%s" alt="">' % e(it["photo"])
        for it in box["items"] if it["photo"].startswith("http"))
    notes = e(doc["notes"]) if doc["notes"] else ""

    return """
<section class="page">
  <div class="boxtag">Inside box packing list &middot; Box %(n)d of %(of)d</div>
  %(head)s
  %(carrier)s
  <table class="lines">
    <thead><tr><th class="c"># of Boxes</th><th>Description</th>
      <th class="t">Tracking</th><th class="q">QTY</th></tr></thead>
    <tbody>%(rows)s</tbody>
    <tfoot><tr><th class="c">Total Boxes</th><td>%(of)d</td>
      <th class="t">Total Quantity</th><td class="q">%(qty)d</td></tr></tfoot>
  </table>
  <div class="notes"><div class="strip">Special Notes / Photos</div>
    <div class="notebody">%(notes)s%(photos)s</div></div>
</section>""" % {
        "n": box["num"], "of": box["of"],
        "head": _head_block(doc, e), "carrier": _carrier_block(doc, e),
        "rows": "".join(rows), "qty": box["qty"],
        "notes": ('<div class="notetext">%s</div>' % notes) if notes else "",
        "photos": ('<div class="shots">%s</div>' % photos) if photos else "",
    }


def _label(doc, box, e):
    """The Outside Box label for one box."""
    ship = e(doc["ship_to"]).replace("\n", "<br>") if doc["ship_to"] else \
        '<span class="hole">ship-to address missing</span>'
    return """
<section class="page label">
  <div class="boxtag">Outside box label</div>
  <div class="lbl-head">%(logo)s
    <div class="lbl-box"><span>BOX</span><b>%(n)d</b><i>of %(of)d</i></div></div>
  <div class="lbl-to"><div class="cap">SHIP TO</div>
    <div class="addr big">%(who)s%(ship)s</div></div>
  <table class="lbl-meta">
    <tr><th>PO Reference</th><td>%(po)s</td></tr>
    <tr><th>Tracking</th><td class="mono">%(tr)s</td></tr>
    <tr><th>Carrier</th><td>%(carrier)s</td></tr>
    <tr><th>Qty in this box</th><td>%(qty)d</td></tr>
  </table>
  <div class="lbl-from">From: %(company)s &middot; %(coline)s</div>
</section>""" % {
        "n": box["num"], "of": box["of"], "logo": _mark("big"),
        "who": ('<b>%s</b><br>' % e(doc["client"])) if doc["client"] else "",
        "ship": ship,
        "po": e(doc["po_ref"]) or "&mdash;",
        "tr": e(doc["tracking"]) or "&mdash;",
        "carrier": e(doc["carrier"]) or "&mdash;",
        "qty": box["qty"],
        "company": e(COMPANY), "coline": e(COMPANY_LINES[0]),
    }


def render(doc, viewer="", token="", box=None):
    """The printable set: a label and a packing list for every box.

    `box` limits it to one box; the default generates all of them, page-broken
    so a single Cmd-P prints the whole shipment.
    """
    e = html.escape
    chosen = [b for b in doc["boxes"] if box in (None, b["num"])]
    pages = "".join(_label(doc, b, e) + _sheet(doc, b, e) for b in chosen)

    problems = flags(doc)
    if problems:
        items = "".join('<li class="%s">%s</li>' % (p["level"], e(p["text"]))
                        for p in problems)
        stops = sum(1 for p in problems if p["level"] == "stop")
        banner = ('<div class="flagbar %s"><b>%d thing%s to check before '
                  'printing</b><ul>%s</ul></div>'
                  % ("bad" if stops else "ok", len(problems),
                     "" if len(problems) == 1 else "s", items))
    else:
        banner = ('<div class="flagbar good"><b>Nothing missing.</b> Every '
                  'field on the template is filled and the quantities match '
                  'the order.</div>')

    xlsx = ""
    if token:
        xlsx = ('<a class="btn" href="?tracking=%s&t=%s&format=xlsx">Export .xlsx</a>'
                % (e(doc["tracking"]), e(token)))

    return """<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Packing list %(tracking)s - %(client)s</title>
<style>
:root{--ink:#1c1c1c;--ink-2:#5a5a5a;--ink-3:#8d8d8d;--line:#c9c9c9;
      --band:%(band)s;--stripe:%(stripe)s;--totals:%(totals)s;
      --red:#b8322b;--amber:#9a6b12;--green:#116a42;--accent:#2440d8}
*{box-sizing:border-box}
html,body{margin:0;background:#e9e8e4}
body{font:13px/1.45 Calibri,-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;
     color:var(--ink);padding:20px 14px 60px;-webkit-font-smoothing:antialiased}
.bar{max-width:900px;margin:0 auto 14px;display:flex;gap:9px;align-items:center;flex-wrap:wrap}
.bar .who{margin-right:auto;color:var(--ink-2);font-size:13px}
.btn{border:1px solid var(--line);background:#fff;color:var(--ink);border-radius:8px;
     padding:8px 15px;font:inherit;font-size:13px;font-weight:600;cursor:pointer;text-decoration:none}
.btn.primary{background:var(--accent);border-color:var(--accent);color:#fff}
.flagbar{max-width:900px;margin:0 auto 16px;background:#fff;border-radius:10px;
         padding:14px 18px;border-left:5px solid var(--amber);
         box-shadow:0 1px 6px rgba(0,0,0,.07);font-size:13px}
.flagbar.bad{border-left-color:var(--red)}
.flagbar.good{border-left-color:var(--green)}
.flagbar b{display:block;margin-bottom:7px;font-size:13.5px}
.flagbar ul{margin:0;padding-left:19px}
.flagbar li{margin:2px 0}
.flagbar li.stop{color:var(--red);font-weight:600}
.flagbar li.warn{color:var(--amber)}
.flagbar li.info{color:var(--ink-3)}
.page{max-width:900px;margin:0 auto 22px;background:#fff;padding:30px 34px 34px;
      box-shadow:0 2px 12px rgba(0,0,0,.08);position:relative}
.boxtag{position:absolute;top:0;left:0;background:var(--band);color:#fff;
        font-size:10px;letter-spacing:.09em;text-transform:uppercase;
        font-weight:700;padding:4px 11px}
.top{display:flex;justify-content:space-between;align-items:flex-start;
     gap:20px;margin-top:16px}
.from{border:1.5px solid #4a6fb5;padding:9px 11px;min-width:340px;line-height:1.5}
.from b{display:block;font-size:14px;letter-spacing:.02em}
.mark{width:118px;height:auto;max-height:88px;object-fit:contain;
      display:block;text-align:center}
div.mark{background:var(--band);color:#fff;height:82px;display:grid;
      place-items:center;font-weight:800;font-size:16px;line-height:1.05;
      letter-spacing:.06em;border-radius:3px}
.mark.big{width:168px;max-height:124px}
div.mark.big{height:104px;font-size:21px}
.mid{display:flex;gap:20px;align-items:flex-start;margin-top:0}
.shipto{flex:1;border:1.5px solid #4a6fb5;border-top:0;min-height:132px;padding:0 11px 11px}
.cap{text-align:center;font-weight:700;text-decoration:underline;
     padding:6px 0 8px;font-size:13px}
.addr{line-height:1.5}
.addr.big{font-size:19px;line-height:1.45;font-weight:600}
.panel{border-collapse:collapse;width:300px}
.panel th{background:#fff;border:1px solid var(--ink);text-align:right;
          padding:5px 9px;font-weight:600;width:130px;font-size:12.5px}
.panel td{border:1px solid var(--ink);padding:5px 9px;min-width:150px}
.carrier{border-collapse:collapse;width:100%%;margin-top:14px;table-layout:fixed}
.carrier th{background:var(--band);color:#fff;font-weight:700;font-size:12.5px;
            border:1px solid var(--band);padding:5px 8px;width:33.33%%}
.carrier td{border:1px solid var(--line);padding:6px 8px;height:26px}
.lines{border-collapse:collapse;width:100%%;margin-top:18px;table-layout:fixed}
.lines th{background:var(--band);color:#fff;font-weight:700;font-size:12.5px;
          border:1px solid var(--band);padding:5px 8px;text-align:left}
.lines td{border:1px solid var(--line);padding:6px 8px;height:25px;
          vertical-align:middle}
.lines tbody tr:nth-child(odd) td{background:var(--stripe)}
.lines tbody tr.blank td{height:25px}
.lines .c{width:92px;text-align:center}
.lines .t{width:190px;font-size:11.5px;word-break:break-all}
.lines .q{width:78px;text-align:right;font-variant-numeric:tabular-nums}
.lines .ord{display:block;color:var(--ink-3);font-size:11px}
.lines tfoot th,.lines tfoot td{background:var(--totals);border:1px solid var(--line);
  color:var(--ink);font-weight:700;text-align:center}
.lines tfoot td.q{text-align:right}
.short{color:var(--red);font-weight:700}
.hole{color:var(--red);font-style:italic}
.notes{margin-top:16px}
.strip{background:%(notes)s;color:#fff;text-align:center;font-weight:700;
       font-size:12.5px;padding:5px}
.notebody{border:1px solid var(--line);border-top:0;min-height:72px;padding:9px}
.notetext{margin-bottom:8px}
.shots{display:flex;gap:8px;flex-wrap:wrap}
.shots img{width:74px;height:74px;object-fit:cover;border:1px solid var(--line);border-radius:3px}
.label .lbl-head{display:flex;justify-content:space-between;align-items:center;margin-top:18px}
.lbl-box{border:3px solid var(--ink);padding:10px 20px;text-align:center;line-height:1}
.lbl-box span{display:block;font-size:11px;letter-spacing:.16em;font-weight:700}
.lbl-box b{display:block;font-size:52px;line-height:1;margin:3px 0}
.lbl-box i{display:block;font-style:normal;font-size:13px;color:var(--ink-2)}
.lbl-to{margin-top:22px;border:2px solid var(--ink);padding:0 16px 18px}
.lbl-meta{border-collapse:collapse;width:100%%;margin-top:16px}
.lbl-meta th{background:var(--band);color:#fff;text-align:left;padding:7px 10px;
             width:190px;font-size:12.5px}
.lbl-meta td{border:1px solid var(--line);padding:7px 10px;font-size:15px;font-weight:600}
.mono{font-family:ui-monospace,SFMono-Regular,Menlo,monospace;font-size:14px}
.lbl-from{margin-top:16px;color:var(--ink-2);font-size:12px}
@media print{
  body{background:#fff;padding:0}
  .bar,.flagbar{display:none}
  .page{box-shadow:none;max-width:none;margin:0;padding:8mm 0 0;
        page-break-after:always;break-after:page}
  .page:last-child{page-break-after:auto;break-after:auto}
  @page{margin:12mm}
}
</style></head><body>
<div class="bar">
  <span class="who">%(count)d box%(es)s &middot; %(client)s &middot; %(tracking)s%(prep)s</span>
  <a class="btn" href="javascript:history.back()">Back</a>
  %(xlsx)s
  <button class="btn primary" onclick="window.print()">Print all</button>
</div>
%(banner)s
%(pages)s
</body></html>""" % {
        "tracking": e(doc["tracking"]), "client": e(doc["client"] or "Shipment"),
        "band": BAND, "stripe": STRIPE, "totals": TOTALS, "notes": NOTES,
        "count": len(chosen), "es": "" if len(chosen) == 1 else "es",
        "prep": (" &middot; prepared by %s" % e(viewer)) if viewer else "",
        "xlsx": xlsx, "banner": banner, "pages": pages,
    }


# ------------------------------------------------------------------- export

def workbook(doc):
    """The same set as .xlsx -- one worksheet per box, template layout."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin", color="FFC9C9C9")
    dark = Side(style="thin", color="FF1C1C1C")
    box_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    box_dark = Border(left=dark, right=dark, top=dark, bottom=dark)
    band = PatternFill("solid", fgColor=BAND_XL)
    stripe = PatternFill("solid", fgColor=STRIPE_XL)
    totals = PatternFill("solid", fgColor=TOTALS_XL)
    white_bold = Font(bold=True, color="FFFFFFFF", size=11)
    bold = Font(bold=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb = Workbook()
    wb.remove(wb.active)

    for box in doc["boxes"]:
        ws = wb.create_sheet("Box %d" % box["num"])
        for col, width in zip("ABCDEF", (14, 24, 22, 6, 26, 12)):
            ws.column_dimensions[col].width = width

        ws.merge_cells("A1:C3")
        ws["A1"] = "%s\n%s\n%s" % (COMPANY, COMPANY_LINES[0], COMPANY_LINES[1])
        ws["A1"].alignment = Alignment(horizontal="left", vertical="top",
                                       wrap_text=True)
        ws["A1"].font = bold
        ws.merge_cells("E1:F3")
        raw = logo_bytes()
        if raw:
            from openpyxl.drawing.image import Image as XLImage
            pic = XLImage(io.BytesIO(raw))
            scale = 84.0 / pic.height
            pic.width, pic.height = int(pic.width * scale), 84
            ws.add_image(pic, "E1")
        else:
            ws["E1"] = COMPANY
            ws["E1"].fill = band
            ws["E1"].font = Font(bold=True, color="FFFFFFFF", size=16)
            ws["E1"].alignment = center

        ws.merge_cells("A4:C4")
        ws["A4"] = "SHIPPING ADDRESS"
        ws["A4"].alignment = center
        ws["A4"].font = Font(bold=True, underline="single")
        ws.merge_cells("A5:C10")
        ws["A5"] = "\n".join(x for x in (doc["client"], doc["ship_to"]) if x)
        ws["A5"].alignment = Alignment(horizontal="left", vertical="top",
                                       wrap_text=True)

        for row, (label, value) in enumerate(
                (("Date:", doc["date_shipped"] or doc["generated"]),
                 ("Invoice #:", doc["invoice_num"]),
                 ("PO Reference:", doc["po_ref"])), start=4):
            ws.cell(row=row, column=5, value=label).alignment = right
            ws.cell(row=row, column=5).border = box_dark
            ws.cell(row=row, column=5).font = bold
            ws.cell(row=row, column=6, value=value).border = box_dark

        for col, label, value in ((1, "Carrier", doc["carrier"]),
                                  (2, "DDP", doc["ddp"]),
                                  (3, "Date Shipped", doc["date_shipped"])):
            h = ws.cell(row=11, column=col, value=label)
            h.fill, h.font, h.alignment, h.border = band, white_bold, center, box_thin
            v = ws.cell(row=12, column=col, value=value)
            v.alignment, v.border = center, box_thin

        for col, label in ((1, "# of Boxes"), (2, "Description"),
                           (5, "Tracking"), (6, "QTY")):
            h = ws.cell(row=14, column=col, value=label)
            h.fill, h.font, h.alignment, h.border = band, white_bold, center, box_thin
        ws.merge_cells("B14:D14")
        for col in (3, 4):
            ws.cell(row=14, column=col).fill = band
            ws.cell(row=14, column=col).border = box_thin

        row = 15
        for n, it in enumerate(box["items"]):
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            ws.cell(row=row, column=1, value=1 if n == 0 else None).alignment = center
            ws.cell(row=row, column=2, value=it["description"]).alignment = left
            ws.cell(row=row, column=5,
                    value=doc["tracking"] if n == 0 else None).alignment = center
            qty = ws.cell(row=row, column=6, value=it["shipped"])
            qty.alignment = right
            if it["short"]:
                qty.font = Font(bold=True, color="FFB8322B")
                ws.cell(row=row, column=6).comment = None
            for col in range(1, 7):
                c = ws.cell(row=row, column=col)
                c.border = box_thin
                if n % 2 == 0:
                    c.fill = stripe
            row += 1

        last = max(row, 20)
        for r in range(row, last):
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
            for col in range(1, 7):
                c = ws.cell(row=r, column=col)
                c.border = box_thin
                if (r - 15) % 2 == 0:
                    c.fill = stripe

        ws.merge_cells(start_row=last, start_column=2, end_row=last, end_column=3)
        ws.cell(row=last, column=2, value="Total Boxes").alignment = center
        ws.cell(row=last, column=4, value=doc["box_count"]).alignment = center
        ws.cell(row=last, column=5, value="Total Quantity").alignment = center
        ws.cell(row=last, column=6, value=box["qty"]).alignment = right
        for col in range(1, 7):
            c = ws.cell(row=last, column=col)
            c.fill, c.font, c.border = totals, bold, box_thin

        note_row = last + 1
        ws.merge_cells(start_row=note_row, start_column=1,
                       end_row=note_row, end_column=6)
        n = ws.cell(row=note_row, column=1, value="Special Notes / Photos")
        n.fill = PatternFill("solid", fgColor="FF1C1C1C")
        n.font = Font(bold=True, color="FFFFFFFF")
        n.alignment = center
        if doc["notes"]:
            ws.merge_cells(start_row=note_row + 1, start_column=1,
                           end_row=note_row + 3, end_column=6)
            ws.cell(row=note_row + 1, column=1, value=doc["notes"]).alignment = left

        ws.print_area = "A1:F%d" % (note_row + 4)
        ws.page_setup.orientation = "portrait"
        ws.page_setup.fitToWidth = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True

    # A sheet naming everything the template couldn't fill, so the workbook
    # carries its own flags rather than losing them on the way out.
    problems = flags(doc)
    if problems:
        ws = wb.create_sheet("Flags")
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 76
        ws["A1"], ws["B1"] = "Level", "What to check before this ships"
        ws["A1"].font = ws["B1"].font = bold
        for i, p in enumerate(problems, start=2):
            ws.cell(row=i, column=1, value=p["level"].upper())
            ws.cell(row=i, column=2, value=p["text"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def filename(doc, ext="xlsx"):
    stem = (doc["po_ref"] or doc["tracking"] or "packing-list").replace("/", "-")
    client = (doc["client"] or "").replace("/", "-").replace(" ", "-")
    return "Packing-List_%s%s.%s" % (client + "_" if client else "", stem, ext)
