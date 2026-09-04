"""Packing lists: gathering, box splitting, flagging, the page and the xlsx."""

import io
import zipfile

import packing_list


PASS = []
FAIL = []


def check(label, cond):
    (PASS if cond else FAIL).append(label)


# A three-line shipment: tracking/carrier/boxes typed once, carried down.
ROWS = [
    {"tracking_num": "1ZC228W90429665565", "carrier": "UPS", "num_boxes": "2",
     "customer": "7Brew Coffee", "shipment_id": "SO-DW-0017", "tab": "Hannah",
     "order_num": "PO-1191", "product_name": "Brewista Crew - Black",
     "product_photo": "https://example.com/black.jpg", "box_num": "1",
     "qty_shipped": "300", "qty_expected": "300", "notes": "Handle with care"},
    {"tracking_num": "1ZC228W90429665565", "order_num": "PO-1191",
     "product_name": "Brewista Crew - Natural", "product_photo": "",
     "box_num": "1", "qty_shipped": "225", "qty_expected": "300"},
    {"tracking_num": "1ZC228W90429665565", "product_name": "Trucker Cap - Red",
     "box_num": "2", "qty_shipped": "200", "qty_expected": "200"},
    # A carry-forward row with nothing on it -- not a line anyone packed.
    {"tracking_num": "1ZC228W90429665565", "product_name": "",
     "qty_shipped": "", "qty_expected": ""},
    # A different shipment entirely.
    {"tracking_num": "794657812345", "carrier": "FEDEX", "customer": "Craftworks",
     "product_name": "Tote - Canvas", "qty_shipped": "50", "qty_expected": "50"},
]

doc = packing_list.gather(ROWS, "1ZC228W90429665565")

# ------------------------------------------------------------- gathering
check("gathers only that tracking number", doc is not None and len(doc["items"]) == 3)
check("skips the empty carry-forward row",
      all(i["description"] for i in doc["items"]))
check("carrier comes off the head row", doc["carrier"] == "UPS")
check("client comes off the head row", doc["client"] == "7Brew Coffee")
check("PO reference comes off the head row", doc["po_ref"] == "PO-1191")
check("notes come off the head row", doc["notes"] == "Handle with care")
check("shipped total", doc["shipped_total"] == 725)
check("ordered total", doc["ordered_total"] == 800)
check("total short", doc["short_total"] == 75)
check("per-line short only on the short line",
      [i["short"] for i in doc["items"]] == [0, 75, 0])
check("tracking match is case-insensitive",
      packing_list.gather(ROWS, "1zc228w90429665565") is not None)
check("unknown tracking is None", packing_list.gather(ROWS, "NOPE-9") is None)
check("blank tracking is None", packing_list.gather(ROWS, "") is None)
check("other shipment gathers on its own",
      len(packing_list.gather(ROWS, "794657812345")["items"]) == 1)

# ----------------------------------------------------------------- boxes
check("splits into the sheet's box count", len(doc["boxes"]) == 2)
check("box 1 holds its two lines", len(doc["boxes"][0]["items"]) == 2)
check("box 2 holds its one line", len(doc["boxes"][1]["items"]) == 1)
check("box quantities are per box",
      [b["qty"] for b in doc["boxes"]] == [525, 200])
check("boxes know the set size", all(b["of"] == 2 for b in doc["boxes"]))
check("allocation is recognised", doc["allocated"] is True)

unallocated = packing_list.gather(
    [{"tracking_num": "U1234567", "num_boxes": "3", "product_name": "Mug",
      "qty_shipped": "10", "qty_expected": "10"}], "U1234567")
check("no allocation puts everything in box 1",
      unallocated["boxes"][0]["qty"] == 10 and unallocated["allocated"] is False)
check("no allocation still makes every box", len(unallocated["boxes"]) == 3)

noboxes = packing_list.gather(
    [{"tracking_num": "N1234567", "product_name": "Mug", "qty_shipped": "5",
      "qty_expected": "5"}], "N1234567")
check("a blank box count means one box", noboxes["box_count"] == 1)

# ----------------------------------------------------------------- flags
f = packing_list.flags(doc)
text = " | ".join(p["text"] for p in f)
check("flags the shortage", "75 units short" in text)
check("shortage is a stop", any(p["level"] == "stop" and "short" in p["text"]
                                for p in f))
check("flags the missing ship-to", "Ship-to address is blank" in text)
check("flags the missing invoice number", "Invoice # is blank" in text)
check("flags the missing DDP", "DDP is blank" in text)
check("flags the missing date shipped", "Date Shipped is blank" in text)
check("does not flag fields that are filled",
      "Carrier is blank" not in text and "PO Reference is blank" not in text)
check("counts the missing photos", "2 lines with no product photo" in text)
check("photos are only info", all(p["level"] == "info" for p in f
                                  if "photo" in p["text"]))
check("flags unallocated boxes",
      any("No box allocation" in p["text"] and p["level"] == "stop"
          for p in packing_list.flags(unallocated)))
check("flags empty boxes",
      any("Box 2 is empty" in p["text"] or "Boxes 2, 3 are empty" in p["text"]
          for p in packing_list.flags(unallocated)))

clean = packing_list.gather(
    [{"tracking_num": "C1234567", "carrier": "UPS", "num_boxes": "1",
      "customer": "7Brew", "ship_to": "1 Main St", "invoice_num": "INV-9",
      "order_num": "PO-2", "tariff_charge": "DDP", "ship_date": "Sep 4, 2026",
      "product_name": "Mug", "product_photo": "https://x/y.jpg",
      "box_num": "1", "qty_shipped": "10", "qty_expected": "10"}], "C1234567")
check("a complete shipment flags nothing", packing_list.flags(clean) == [])

# ------------------------------------------------------------------ page
page = packing_list.render(doc, viewer="Hannah", token="tok123")
for needle in ("OFF MENU", "115 Monterey Bay Drive, Boynton Beach, FL 33426",
               "www.byoffmenu.com", "SHIPPING ADDRESS", "Date:", "Invoice #:",
               "PO Reference:", "Carrier", "DDP", "Date Shipped",
               "# of Boxes", "Description", "Tracking", "QTY",
               "Total Boxes", "Total Quantity", "Special Notes / Photos",
               "7Brew Coffee", "1ZC228W90429665565", "UPS",
               "Brewista Crew - Natural", "Handle with care",
               "Outside box label", "Box 1 of 2", "Box 2 of 2",
               "window.print()", "format=xlsx"):
    check("page shows %r" % needle, needle in page)
check("page shows the shortage inline", "225 of 300" in page)
check("page carries the flag banner", "to check before printing" in page)
check("photo renders as an image", 'src="https://example.com/black.jpg"' in page)
check("blank fields are marked as holes", 'class="hole"' in page)
check("pages break for printing", "page-break-after:always" in page)
check("a clean shipment says nothing is missing",
      "Nothing missing." in packing_list.render(clean))
check("no token means no export button",
      "format=xlsx" not in packing_list.render(doc))

one = packing_list.render(doc, box=2)
check("?box= narrows to one box", "Box 2 of 2" in one and "Box 1 of 2" not in one)

nasty = packing_list.gather(
    [{"tracking_num": "X1234567", "customer": "<script>alert(1)</script>",
      "product_name": "Cap & Co", "qty_shipped": "1", "qty_expected": "1"}],
    "X1234567")
out = packing_list.render(nasty)
check("user text is escaped", "<script>" not in out and "&lt;script&gt;" in out)
check("ampersands survive escaping", "Cap &amp; Co" in out)

qty = packing_list.gather(
    [{"tracking_num": "Y1234567", "product_name": "Mug",
      "qty_shipped": "1,200", "qty_expected": "1200.0"}], "Y1234567")
check("commas and decimals parse",
      qty["shipped_total"] == 1200 and qty["ordered_total"] == 1200)

# ------------------------------------------------------------------ xlsx
blob = packing_list.workbook(doc)
check("workbook is a real xlsx", blob[:2] == b"PK" and len(blob) > 4000)
names = zipfile.ZipFile(io.BytesIO(blob)).namelist()
check("workbook has worksheets", any("worksheets/sheet" in n for n in names))

from openpyxl import load_workbook
wb = load_workbook(io.BytesIO(blob))
check("a worksheet per box", "Box 1" in wb.sheetnames and "Box 2" in wb.sheetnames)
check("flags ride along in the workbook", "Flags" in wb.sheetnames)
ws = wb["Box 1"]
check("xlsx carries the letterhead", "OFF MENU" in str(ws["A1"].value))
check("xlsx carries the section headers",
      ws["A4"].value == "SHIPPING ADDRESS" and ws["A11"].value == "Carrier"
      and ws["A14"].value == "# of Boxes")
check("xlsx carries the client", "7Brew Coffee" in str(ws["A5"].value))
check("xlsx carries the PO reference", ws["F6"].value == "PO-1191")
check("xlsx carries the line items",
      ws["B15"].value == "Brewista Crew - Black" and ws["F15"].value == 300)
check("xlsx carries the tracking once", ws["E15"].value == "1ZC228W90429665565"
      and ws["E16"].value is None)
check("xlsx box 2 holds its own line",
      wb["Box 2"]["B15"].value == "Trucker Cap - Red")
check("xlsx flags sheet lists the problems",
      any("short" in str(c.value or "") for c in wb["Flags"]["B"]))
check("filename names the client and PO",
      packing_list.filename(doc) == "Packing-List_7Brew-Coffee_PO-1191.xlsx")

print("\n".join("  ok   %s" % p for p in PASS))
if FAIL:
    print("\n".join("  FAIL %s" % f for f in FAIL))
print("\n%d passed, %d failed" % (len(PASS), len(FAIL)))
raise SystemExit(1 if FAIL else 0)
